# -*- coding: utf-8 -*-
import glob
import os
import json
import datetime
import pytz
import re
import csv
from openpyxl import load_workbook


BUY_PATH = "/home/nigam_parikh_scp/FM/results/Bookscouter/Trade-In/Buy_Pricing/"
SELL_PATH = "/home/nigam_parikh_scp/FM/results/Bookscouter/Trade-In/Sell_Pricing/"

COMBINE_PATH = "/home/nigam_parikh_scp/FM/results/Combine"

# BUY_PATH = '/Users/PathakUmesh/Programming_stuffs/NIGAM/BUYING/*'
# SELL_PATH = '/Users/PathakUmesh/Programming_stuffs/NIGAM/SELLING/*'
# COMBINE_PATH = '/Users/PathakUmesh/Programming_stuffs/NIGAM/COMBINE'




def get_file(path):
    list_of_files = glob.glob(path)
    file_path = max(list_of_files, key=os.path.getctime)
    return file_path

def combine_buy_and_sell():
    buy_file_name = get_file(BUY_PATH)
    book = load_workbook(buy_file_name)
    sheet = book.active

    print('---------------------------------------------------------------------------------------------------------------------')
    print('Taking buying details from file {}'.format(buy_file_name))
    print('---------------------------------------------------------------------------------------------------------------------')
    
    sell_file_name = get_file(SELL_PATH)

    print('---------------------------------------------------------------------------------------------------------------------')
    print('Taking selling details from file {}'.format(sell_file_name))
    print('---------------------------------------------------------------------------------------------------------------------')
    

    if not os.path.exists(COMBINE_PATH):
        os.makedirs(COMBINE_PATH)
    
    utc_time = datetime.datetime.utcnow()
    tz_info = pytz.timezone('Asia/Kolkata')
    utc = pytz.utc
    time_local = utc.localize(utc_time).astimezone(tz_info)
    
    combine_file_name = '{}/Combine_Trade-In_{}.csv'.format(COMBINE_PATH,time_local.strftime('%d%b%Y_%Hhr%Mmin'))
    combine_filtered_file_name = '{}/Combine_Trade-In_Filtered_{}.csv'.format(COMBINE_PATH,time_local.strftime('%d%b%Y_%Hhr%Mmin'))
    
    combine_csvfile =  open(combine_file_name, 'w')
    combine_filtered_csvfile =  open(combine_filtered_file_name, 'w')
    
    fieldnames = ['Title', 'ISBN-10', 'ISBN-13', 'URL_Buy', 'Condition', 'Description',
                  'Purchase Cost', 'Shipping Cost', 'Processing Cost', 'Net Buying Cost',
                  'Vendor', 'Selling Price','ROI Amt', 'ROI %']
    writer1 = csv.DictWriter(combine_csvfile, fieldnames=fieldnames)
    writer1.writeheader()

    writer2 = csv.DictWriter(combine_filtered_csvfile, fieldnames=fieldnames)
    writer2.writeheader()
    print('---------------------------------------------------------------------------------------------------------------------')
    print('Writing output to file {}'.format(combine_file_name))
    print('---------------------------------------------------------------------------------------------------------------------')

    print('---------------------------------------------------------------------------------------------------------------------')
    print('Writing filtered output to file {}'.format(combine_filtered_file_name))
    print('---------------------------------------------------------------------------------------------------------------------')




    for i, buy_row in enumerate(sheet.iter_rows('B{}:K{}'.format(sheet.min_row,sheet.max_row))):
        if not i == 0:
            with open(sell_file_name, 'r') as csvfile:
                csvreader = csv.reader(csvfile)
                for j, sell_row in enumerate(csvreader):
                    if not j == 0:
                        if str(buy_row[1].value) == sell_row[1]:
                            writer1.writerow({'Title': str(buy_row[0].value),
                                'ISBN-10': str(buy_row[1].value),
                                'ISBN-13': str(buy_row[2].value),
                                'URL_Buy': str(buy_row[3].value),
                                'Condition': str(buy_row[4].value),
                                'Description': str(buy_row[5].value),
                                'Purchase Cost': str(buy_row[6].value),
                                'Shipping Cost': str(buy_row[7].value),
                                'Processing Cost': str(buy_row[8].value),
                                'Net Buying Cost': str(buy_row[9].value),
                                'Vendor': sell_row[3],
                                'Selling Price': sell_row[4],
                                'ROI Amt': '%.2f' % (float(sell_row[4]) - float(buy_row[9].value)),
                                'ROI %': '%.2f' % ((float(sell_row[4]) - float(buy_row[9].value))* 100/float(buy_row[9].value))
                            })
                            if (float(sell_row[4]) - float(buy_row[9].value)) >= 10:
                                writer2.writerow({'Title': str(buy_row[0].value),
                                    'ISBN-10': str(buy_row[1].value),
                                    'ISBN-13': str(buy_row[2].value),
                                    'URL_Buy': str(buy_row[3].value),
                                    'Condition': str(buy_row[4].value),
                                    'Description': str(buy_row[5].value),
                                    'Purchase Cost': str(buy_row[6].value),
                                    'Shipping Cost': str(buy_row[7].value),
                                    'Processing Cost': str(buy_row[8].value),
                                    'Net Buying Cost': str(buy_row[9].value),
                                    'Vendor': sell_row[3],
                                    'Selling Price': sell_row[4],
                                    'ROI Amt': '%.2f' % (float(sell_row[4]) - float(buy_row[9].value)),
                                    'ROI %': '%.2f' % ((float(sell_row[4]) - float(buy_row[9].value))* 100/float(buy_row[9].value))
                                })


if __name__ == '__main__':
    combine_buy_and_sell()