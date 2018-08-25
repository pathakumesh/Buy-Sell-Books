# -*- coding: utf-8 -*-
import glob
import os
import json
import scrapy
import datetime
import pytz
from scrapy.crawler import CrawlerProcess
from openpyxl import load_workbook
import logging as log
# log.getLogger('scrapy').propagate = False
# log.getLogger('rotating_proxies').propagate = False

# INPUT_PATH = "/home/nigam_parikh_scp/FM/results/Amazon/Trade-In/ISBN/*"
INPUT_PATH = '/Users/PathakUmesh/Programming_stuffs/NIGAM/ISBN/*'

class ExtractItem(scrapy.Item):
    # define the fields for your item here like:
    title = scrapy.Field()
    isbn10 = scrapy.Field()
    isbn13 = scrapy.Field()
    condition = scrapy.Field()
    vendor = scrapy.Field()
    purchase_cost = scrapy.Field()
    shipping_cost = scrapy.Field()
    processing_cost = scrapy.Field()
    net_buying_cost = scrapy.Field()
    url_buy = scrapy.Field()
    pass    

class BookScouterSellSpider(scrapy.Spider):
    name = "book_scouter_sell_spider"
    allowed_domains = ["bookscouter.com"]

    def __init__(self, cookie=None):
        self.cookie = cookie
    
    def read_isbns(self):
        list_of_files = glob.glob(INPUT_PATH)
        file_path = max(list_of_files, key=os.path.getctime)
        file_path = '/Users/PathakUmesh/Programming_stuffs/NIGAM/ISBN/testing.xlsx'

        log.info('[+++++] Starting from input file path: /home/nigam_parikh_scp/FM/results/Amazon/Trade-In/ISBN/')
        log.info('[+++++] Starting with input file name: {}'.format(file_path.rsplit('/',1)[-1]))
        
        book = load_workbook(file_path)
        sheet = book.active
        isbns = list()
        for row in sheet.iter_rows('D{}:D{}'.format(sheet.min_row,sheet.max_row)):
            isbns.append([str(cell.value) for cell in row][0])
        return isbns[1:]
        
    def start_requests(self):
        isbns = self.read_isbns()
        isbn_url = 'https://api.bookscouter.com/v3/prices/buy/{}'
        print('Cookie is:', self.cookie)
        headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'en-US,en;q=0.9,hi;q=0.8',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Host': 'api.bookscouter.com',
            'Pragma': 'no-cache',
            'Upgrade-Insecure-Requests': 1,
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36',
            'Cookie': self.cookie
        }

        for page_number in range(len(isbns)):
            yield scrapy.Request(
                url= isbn_url.format(isbns[page_number]),
                callback=self.parse,
                headers =headers,
                dont_filter = True
            )

    def parse(self, response):
        json_data = json.loads(response.text)
        log.info('[+++] Status code {} for url {}'.format(response.status, response.url))
        if json_data.get('data') and json_data['data'].get('Book') and json_data['data'].get('Prices'):
            
            book_data = json_data['data']['Book']
            
            title = book_data['Title']
            isbn10 = book_data['Isbn10']
            isbn13 = book_data['Isbn13']
            
            prices = json_data['data']['Prices']
            amazon_prices = list(filter(lambda x:x['Vendor']['Name'] == 'Amazon.com' and x['Price'] > 0, prices))
            if amazon_prices:
                # log.info('[+++] Minimum Amazon Price found for url {}'.format(response.url))
                final_price = min(amazon_prices, key=lambda x:x['Price'])
                
                item = ExtractItem()
                purchase_cost = float('%.2f' % final_price['Price'])
                shipping_cost = float('%.2f' % final_price['Shipping'])
                vendor = final_price['Vendor'].get('Name', '')
                condition = final_price['Condition']
                url_buy = final_price['Url']

                item['title'] = title
                item['isbn10'] = isbn10
                item['isbn13'] = isbn13
                item['vendor'] = vendor
                item['purchase_cost'] = purchase_cost
                item['shipping_cost'] = shipping_cost
                item['processing_cost'] = 2
                item['net_buying_cost'] = float('%.2f' % (purchase_cost + shipping_cost + 2))
                item['url_buy'] = url_buy
                item['condition'] = condition
                yield item



def run_spider(no_of_threads, request_delay, cookie):

    log.info('[+++++] Starting with Threads:{} and Delay between Request: {}'.format(no_of_threads, request_delay))

    rotating_proxy_list = 'helper_files/proxy.txt'
    settings = {"DOWNLOADER_MIDDLEWARES":{
                            'scrapy.downloadermiddlewares.retry.RetryMiddleware': 90,
                            'rotating_proxies.middlewares.RotatingProxyMiddleware': 610,
                            'rotating_proxies.middlewares.BanDetectionMiddleware': 620,
                            'scrapy.contrib.downloadermiddleware.cookies.CookiesMiddleware': 700
                },
                'ITEM_PIPELINES':{
                            'pipelines_bookscouter_buy.ExtractPipeline': 300,
                },
                'USER_AGENT': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36',
                'DOWNLOAD_DELAY': request_delay,
                'COOKIES_ENABLED': True,
                'CONCURRENT_REQUESTS':no_of_threads,
                'RETRY_TIMES':10,
                'ROTATING_PROXY_PAGE_RETRY_TIMES':10,
                'RETRY_HTTP_CODES':[403, 429],
                'ROTATING_PROXY_LIST_PATH':rotating_proxy_list,
                'ROTATING_PROXY_BAN_POLICY': 'pipelines_bookscouter_buy.BanPolicy',
                # 'LOG_ENABLED':False

    }
    # -----------------------Run without Proxy----------------------------------------
    # settings = {
    #             'ITEM_PIPELINES':{
    #                         'pipelines.ExtractPipeline': 300,
    #             },
    #             'USER_AGENT': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36',
    #             'DOWNLOAD_DELAY': request_delay,
    #             'RETRY_TIMES':10,
    #             'RETRY_HTTP_CODES':[429],
    #             'LOG_ENABLED':False
    # }

    process = CrawlerProcess(settings)
    process.crawl(BookScouterSellSpider, cookie = cookie)

    process.start()

if __name__ == '__main__':
    no_of_threads = 40
    request_delay = 0.5
    cookie = '_ga=GA1.2.454851917.1533183282; __stripe_mid=0e997e28-5121-4a81-b131-dfd77e646a7c; bsLoadCheck=loaded; _gid=GA1.2.1350737974.1534429184; _gat_UA-542071-3=1; __stripe_sid=5272cace-8a0e-4b51-a994-3e6b2cf03bf0'

    run_spider(no_of_threads, request_delay, cookie)

