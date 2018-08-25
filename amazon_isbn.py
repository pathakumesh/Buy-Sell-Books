# -*- coding: utf-8 -*-
import glob
import os
import csv
import json
import scrapy
import datetime
import pytz
from scrapy.crawler import CrawlerProcess
from openpyxl import load_workbook
import logging as log


INPUT_PATH = "/home/FM/results/Category_Indexing/*"
# INPUT_PATH = '/Users/PathakUmesh/Programming_stuffs/NIGAM/CATEGORY/*'


class ExtractItem(scrapy.Item):
    # define the fields for your item here like:
    title = scrapy.Field()
    isbn10 = scrapy.Field()
    isbn13 = scrapy.Field()
    pass    

class AmazonISBNSpider(scrapy.Spider):
    name = "amazon_isbn_spider"
    allowed_domains = ["amazon.com"]
    extracted_isbns = list()
    # def read_urls(self):
    #     list_of_files = glob.glob(INPUT_PATH)
    #     file_path = max(list_of_files, key=os.path.getctime)
    #     # file_path = '/Users/PathakUmesh/Programming_stuffs/NIGAM/CATEGORY/Amazon_Categories_13Aug2018_12hr34min.xlsx'
        
    #     log.info('[+++++] Starting from input file path: /home/FM/results/Category_Indexing/')
    #     log.info('[+++++] Starting with input file name: {}'.format(file_path.rsplit('/',1)[-1]))
        
    #     book = load_workbook(file_path)
    #     sheet = book.active
    #     urls = list()
    #     for row in sheet.iter_rows('C{}:C{}'.format(sheet.min_row,sheet.max_row)):
    #         urls.append(str(row[0].value))
            
    #     return urls[1:]

    def read_urls(self):
        list_of_files = glob.glob(INPUT_PATH)
        file_path = max(list_of_files, key=os.path.getctime)
        # file_path = '/Users/PathakUmesh/Programming_stuffs/NIGAM/CATEGORY/Amazon_Categories_13Aug2018_12hr34min.xlsx'
        
        log.info('[+++++] Starting from input file path: /home/FM/results/Category_Indexing/')
        log.info('[+++++] Starting with input file name: {}'.format(file_path.rsplit('/',1)[-1]))
        
        urls = list()
        with open(file_path, 'r') as csvfile:
            csvreader = csv.reader(csvfile)
            for i, url_row in enumerate(csvreader):
                if i == 0:
                    continue
                urls.append(str(url_row[1]))            
        return urls
        
    def format_isbn(self, isbn, length):
        prefix = ''
        for i in range(len(isbn), length):
            prefix += '0'
        return prefix + isbn

    def start_requests(self):
        urls = self.read_urls()
        for amazon_url in urls:
            yield scrapy.Request(
                url=amazon_url,
                callback=self.parse,
                dont_filter=True
            )

    def parse(self, response):
        # print(response.text)
        results = response.xpath('//li[contains(@id, "result_")]')
        for result in results:
            title = result.xpath('div//a/h2/text()').extract_first()
            isbn10 = result.xpath('div//span[contains(text(), "ISBN-10")]/../following-sibling::div[1]/span/text()').extract_first()
            isbn13 = result.xpath('div//span[contains(text(), "ISBN-13")]/../following-sibling::div[1]/span/text()').extract_first()
            
            if isbn10 and len(isbn10) < 10:
                isbn10 = self.format_isbn(isbn10, 10)
            
            if isbn13 and len(isbn13) < 13:
                isbn13 = self.format_isbn(isbn13, 13)

            if isbn10 and isbn13:
                item = ExtractItem()
                item['title'] = title
                item['isbn10'] = isbn10
                item['isbn13'] = isbn13
                yield item
            
        next_page_link = response.xpath('//a[@id="pagnNextLink"]/@href').extract_first()
        if next_page_link and 'https://' not in next_page_link:
            next_page_link = "https://www.amazon.com" + next_page_link
            yield scrapy.Request(
                url= next_page_link,
                callback=self.parse,
                dont_filter = True
            )

def run_spider(no_of_threads, request_delay, timeout, no_of_retries):

    log.info('[+++++] Starting with Threads:{}, Timeout:{}, Retries:{} and Delay between Request: {}'
            .format(no_of_threads, timeout,no_of_retries,request_delay)
            )

    rotating_proxy_list = 'helper_files/proxy.txt'
    settings = {"DOWNLOADER_MIDDLEWARES":{
                            'scrapy.downloadermiddlewares.useragent.UserAgentMiddleware': None,
                            'scrapy_fake_useragent.middleware.RandomUserAgentMiddleware': 400,
                            'scrapy.downloadermiddlewares.retry.RetryMiddleware': 90,
                            'rotating_proxies.middlewares.RotatingProxyMiddleware': 610,
                            'rotating_proxies.middlewares.BanDetectionMiddleware': 620,
                },
                'ITEM_PIPELINES':{
                            'pipelines_amazon_isbn.ExtractPipeline': 300,
                },
                'DOWNLOAD_DELAY': request_delay,
                'DOWNLOAD_TIMEOUT': timeout,
                'CONCURRENT_REQUESTS':no_of_threads,
                'CONCURRENT_REQUESTS_PER_DOMAIN': no_of_threads,
                'RETRY_TIMES':no_of_retries,
                'RANDOM_UA_PER_PROXY': True,
                'ROTATING_PROXY_PAGE_RETRY_TIMES':no_of_retries,
                # 'RETRY_HTTP_CODES':[403, 429],
                'ROTATING_PROXY_LIST_PATH':rotating_proxy_list,
                'ROTATING_PROXY_BAN_POLICY': 'pipelines_amazon_isbn.BanPolicy',
                'LOG_ENABLED':False,

    }
    # -----------------------Run without Proxy----------------------------------------
    # settings = {
    #             'ITEM_PIPELINES':{
    #                         'pipelines.ExtractPipeline': 300,
    #             },
    #             'USER_AGENT': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36',
    #             'DOWNLOAD_DELAY': request_delay,
    #             'RETRY_TIMES':10,
    #             'RETRY_HTTP_CODES':[429],
    #             'LOG_ENABLED':False
    # }

    process = CrawlerProcess(settings)
    process.crawl(AmazonISBNSpider)

    process.start()

if __name__ == '__main__':
    no_of_threads = 80
    request_delay = 0.01
    timeout = 45
    no_of_retries = 10
    run_spider(no_of_threads, request_delay, timeout, no_of_retries)

